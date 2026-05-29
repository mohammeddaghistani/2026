import os
import logging
import tempfile
from pathlib import Path
from io import BytesIO

from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)
from telegram import InputFile

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

DATA_DIR = Path("downloads")
DATA_DIR.mkdir(exist_ok=True)

from ocr import extract_text_from_pdf, extract_text_from_image
from extractor import extract_data
from qreader import read_qr_from_image, read_qr_from_pdf, extract_ticket_from_qr, HAS_QR
from db import init_db, save_extraction, get_history, get_stats, get_pilgrims_by_flight, get_pilgrims_by_airline, passport_exists, lookup_passport_registry
from lang import t, label, START_MSG, DIV
from sheets import export_to_sheets

init_db()

user_data_store = {}
ALLOWED_EXTS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp", ".webp"}

# ── Registration conversation states ──
REG_NAME, REG_PASSPORT, REG_DEP_DATE, REG_DEP_TIME, REG_FLIGHT, REG_LOCATION, REG_FILE_TICKET, REG_FILE_PASSPORT, REG_DECLARATION = range(9)

LOCATION_KEYBOARD = InlineKeyboardMarkup([
    [InlineKeyboardButton("✈️ Jeddah Airport (KAIA)", callback_data="loc:Jeddah Airport (KAIA)")],
    [InlineKeyboardButton("✈️ Medina Airport", callback_data="loc:Medina Airport")],
    [InlineKeyboardButton("🛃 Land Border", callback_data="loc:Land Border")],
])

AGREE_KEYBOARD = InlineKeyboardMarkup([
    [InlineKeyboardButton("✅ I Agree / أوافق", callback_data="agree:yes")],
])


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(START_MSG)


async def process_file(
    update: Update, context: ContextTypes.DEFAULT_TYPE, msg,
    file_bytes: bytes, file_name: str = "file"
) -> None:
    ext = Path(file_name).suffix.lower()
    if ext not in ALLOWED_EXTS:
        await msg.edit_text(t("unsupported"))
        return

    try:
        text = ""
        qr_data = []

        if ext == ".pdf":
            text = extract_text_from_pdf(file_bytes)
            if HAS_QR:
                qr_data = read_qr_from_pdf(file_bytes)
        else:
            text = extract_text_from_image(file_bytes)
            if HAS_QR:
                qr_data = read_qr_from_image(file_bytes)

    except Exception as e:
        logger.exception("Error processing file")
        await msg.edit_text(t("error", str(e)))
        return

    data = extract_data(text)

    if not data["pilgrims"] and not data["tickets"]:
        await msg.edit_text(t("no_text"))
        return

    user_id = update.effective_user.id
    username = update.effective_user.username or ""
    user_data_store[user_id] = data

    save_extraction(
        user_id=user_id,
        username=username,
        file_name=file_name,
        file_type=ext,
        pilgrims=data["pilgrims"],
        tickets=data["tickets"],
        raw_text=data["raw_text"],
    )

    pilgrims = data.get("pilgrims", [])
    tickets = data.get("tickets", {})
    raw = data.get("raw_text", "")

    lines = []
    if pilgrims:
        lines.append("👥 *Pilgrims*")
        for i, p in enumerate(pilgrims, 1):
            lines.append(f"  {i}. {p['name']}")
        lines.append(DIV)

    if tickets:
        lines.append("🎫 *Ticket Details*")
        eng_labels = {
            "flight_number": "Flight", "ticket_number": "Ticket", "seat": "Seat",
            "date": "Date", "gate": "Gate", "airline": "Airline", "passport": "Passport",
            "booking": "Booking", "class": "Class", "destination": "To",
            "origin": "From", "route": "Route", "departure_time": "Departure Time",
            "flight_date": "Flight Date", "hid": "HID",
        }
        for k, v in tickets.items():
            lbl = eng_labels.get(k, k.replace("_", " ").title())
            lines.append(f"  • {lbl}: `{v}`")
        lines.append(DIV)

    if qr_data:
        lines.append("📱 *QR Data*")
        for q in qr_data[:3]:
            lines.append(f"  `{q[:100]}`")
        lines.append(DIV)

    lines.append("📝 *Raw Text*")
    lines.append(f"`{raw[:300]}`")
    if len(raw) > 300:
        lines.append("  ...")

    await msg.edit_text("\n".join(lines))

    if pilgrims or tickets:
        await update.message.reply_text(t("export_prompt"))


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = await update.message.reply_text(t("processing"))
    file = await update.message.document.get_file()
    file_bytes = await file.download_as_bytearray()
    file_name = update.message.document.file_name or "file.pdf"
    await process_file(update, context, msg, bytes(file_bytes), file_name)


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = await update.message.reply_text(t("processing"))
    file = await update.message.photo[-1].get_file()
    file_bytes = await file.download_as_bytearray()
    await process_file(update, context, msg, bytes(file_bytes), "photo.jpg")


async def handle_media_group(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message.media_group_id:
        return
    msgs = context.user_data.get("media_group", [])
    msgs.append(update.message)
    context.user_data["media_group"] = msgs

    if update.message.document:
        msg = await update.message.reply_text(t("processing"))
        file = await update.message.document.get_file()
        fb = await file.download_as_bytearray()
        fn = update.message.document.file_name or "file.pdf"
        await process_file(update, context, msg, bytes(fb), fn)
    elif update.message.photo:
        msg = await update.message.reply_text(t("processing"))
        file = await update.message.photo[-1].get_file()
        fb = await file.download_as_bytearray()
        await process_file(update, context, msg, bytes(fb), "photo.jpg")


async def export_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    data = user_data_store.get(user_id)

    if not data:
        await update.message.reply_text(t("no_data"))
        return

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        await update.message.reply_text("openpyxl not installed")
        return

    wb = openpyxl.Workbook()
    pilgrims = data.get("pilgrims", [])
    tickets = data.get("tickets", {})
    raw = data.get("raw_text", "")

    ws1 = wb.active
    ws1.title = "Pilgrims"
    ws1.append(["#", "Name", "Flight", "Ticket", "Seat", "Airline"])
    for c in "ABCDEF":
        ws1[f"{c}1"].font = Font(bold=True)
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 20

    for i, p in enumerate(pilgrims, 1):
        ws1.append([
            i, p["name"],
            tickets.get("flight_number", ""),
            tickets.get("ticket_number", ""),
            tickets.get("seat", ""),
            tickets.get("airline", ""),
        ])

    ws2 = wb.create_sheet("Raw Text")
    ws2.column_dimensions["A"].width = 120
    for line in raw.split("\n"):
        ws2.append([line])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(DATA_DIR))
    wb.save(tmp.name)
    tmp.close()

    with open(tmp.name, "rb") as f:
        await update.message.reply_document(
            document=InputFile(f, filename="pilgrims_data.xlsx"),
            caption=t("export_ok"),
        )
    os.unlink(tmp.name)


async def history_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    history = get_history(user_id)
    if not history:
        await update.message.reply_text(t("no_data"))
        return

    total = len(history)
    text_lines = [t("history", str(total)) + "\n"]
    for h in history:
        names = ", ".join(p["name"] for p in h["pilgrims"][:3])
        if len(h["pilgrims"]) > 3:
            names += " ..."
        when = h["created_at"][:19].replace("T", " ")
        text_lines.append(f"🆔 {h['id']} | 🕐 {when}")
        text_lines.append(f"  📄 {h['file_name'] or h['file_type']}: {names}")
        text_lines.append("")

    full_text = "\n".join(text_lines)

    if len(full_text) > 4000:
        tmp = tempfile.NamedTemporaryFile(
            suffix=".txt", delete=False, mode="w", dir=str(DATA_DIR)
        )
        tmp.write(full_text)
        tmp.close()
        with open(tmp.name, "rb") as f:
            await update.message.reply_document(
                document=InputFile(f, filename="history.txt"),
                caption=t("history", str(total)),
            )
        os.unlink(tmp.name)
    else:
        await update.message.reply_text(full_text)


async def stats_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    stats = get_stats(user_id)

    if stats["total_pilgrims"] == 0:
        await update.message.reply_text(t("no_data"))
        return

    lines = [f"**{t('stats')}:**\n"]
    lines.append(f"👥 {t('pilgrims')}: {stats['total_pilgrims']}")
    lines.append(f"📄 Files: {stats['total_files']}")
    lines.append(f"✈️ Flights: {stats['total_flights']}")
    lines.append(f"🏢 Airlines: {stats['total_airlines']}")

    flights = get_pilgrims_by_flight(user_id)
    if flights:
        lines.append(f"\n**{t('group_flight')}:**")
        for f in flights[:5]:
            lines.append(f"  • {f['flight_number']}: {f['count']} pilgrims")

    airlines = get_pilgrims_by_airline(user_id)
    if airlines:
        lines.append(f"\n**{t('group_airline')}:**")
        for a in airlines[:5]:
            lines.append(f"  • {a['airline']}: {a['count']} pilgrims")

    await update.message.reply_text("\n".join(lines))


async def exportall_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    history = get_history(user_id)
    if not history:
        await update.message.reply_text(t("no_data"))
        return

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        await update.message.reply_text("openpyxl not installed")
        return

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "All Extractions"
    ws1.append(["#", "Date/Time", "File", "Pilgrims", "Flight", "Ticket", "Seat", "Airline", "Passport"])
    for c in "ABCDEFGHI":
        ws1[f"{c}1"].font = Font(bold=True)
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 40
    ws1.column_dimensions["E"].width = 15
    ws1.column_dimensions["F"].width = 18
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 20
    ws1.column_dimensions["I"].width = 18

    row = 1
    for h in history:
        pilgrims = h.get("pilgrims", [])
        tickets = h.get("tickets", {})
        when = h["created_at"][:19].replace("T", " ")
        pilgrim_names = "; ".join(p["name"] for p in pilgrims)
        ws1.append([
            h["id"], when, h.get("file_name", h.get("file_type", "")),
            pilgrim_names,
            tickets.get("flight_number", ""),
            tickets.get("ticket_number", ""),
            tickets.get("seat", ""),
            tickets.get("airline", ""),
            tickets.get("passport", ""),
        ])
        row += 1

    ws1.auto_filter.ref = f"A1:I{row}"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=str(DATA_DIR))
    wb.save(tmp.name)
    tmp.close()

    with open(tmp.name, "rb") as f:
        await update.message.reply_document(
            document=InputFile(f, filename="all_extractions.xlsx"),
            caption=f"{t('export_ok')} | {t('history', str(len(history)))}",
        )
    os.unlink(tmp.name)


async def sheets_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    data = user_data_store.get(user_id)
    if not data:
        await update.message.reply_text(t("no_data"))
        return

    result = export_to_sheets(data["pilgrims"], data["tickets"])
    await update.message.reply_text(result)


# ────────────────────── Registration Conversation ──────────────────────

async def register_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["reg"] = {}
    await update.message.reply_text(
        "📋 *تسجيل حاج جديد | New Pilgrim Registration*\n"
        "─" * 20 + "\n\n"
        "👤 *1/8* الاسم الكامل للحاج (كما هو في الجواز)\n"
        "Traveler's Full Name (As written in passport)\n\n"
        "أرسل الاسم | Send the name:"
    )
    return REG_NAME


async def reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["reg"]["name"] = update.message.text.strip()
    await update.message.reply_text(
        "🛂 *2/8* رقم جواز السفر\n"
        "Passport Number\n\n"
        "أرسل رقم الجواز (بدون مسافات) | Send passport number (no spaces):"
    )
    return REG_PASSPORT


async def reg_passport(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    passport = update.message.text.strip()
    if " " in passport:
        await update.message.reply_text("❌ رقم الجواز لا يقبل مسافات | No spaces allowed\nأعد الإرسال | Send again:")
        return REG_PASSPORT
    if passport_exists(passport):
        await update.message.reply_text(
            "❌ هذا الجواز مسجل مسبقاً!\n"
            "This passport is already registered!\n\n"
            "للاستفسار تواصل مع الإدارة | Contact admin for help.\n"
            "أرسل /register لتجربة جواز آخر | Send /register to try another passport."
        )
        return ConversationHandler.END
    context.user_data["reg"]["passport"] = passport
    registry_entry = lookup_passport_registry(passport)
    name_note = ""
    if registry_entry and registry_entry.get("name"):
        context.user_data["reg"]["name"] = registry_entry["name"]
        name_note = (
            f"\n\n📋 *Found in registry:* {registry_entry['name']}\n"
            f"سيتم استخدام هذا الاسم — Name set from registry ✅"
        )
    await update.message.reply_text(
        "✅ تم حفظ الجواز | Passport saved\n─" * 15 + "\n\n"
        "📅 *3/8* تاريخ المغادرة\n"
        "Departure Date\n\n"
        "أرسل التاريخ بصيغة YYYY-MM-DD أو اكتب (تخطي/skip)\n"
        "Send date as YYYY-MM-DD or type (skip):" + name_note
    )
    return REG_DEP_DATE


async def reg_dep_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text.lower() in ("skip", "تخطي"):
        context.user_data["reg"]["departure_date"] = ""
    else:
        context.user_data["reg"]["departure_date"] = text
    await update.message.reply_text(
        "⏰ *4/8* وقت المغادرة (حسب التذكرة)\n"
        "Departure Time (As per ticket)\n\n"
        "أرسل الوقت بصيغة HH:MM أو اكتب (تخطي/skip)\n"
        "Send time as HH:MM or type (skip):"
    )
    return REG_DEP_TIME


async def reg_dep_time(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text.lower() in ("skip", "تخطي"):
        context.user_data["reg"]["departure_time"] = ""
    else:
        context.user_data["reg"]["departure_time"] = text
    await update.message.reply_text(
        "✈️ *5/8* رقم الرحلة / الحافلة\n"
        "Flight / Bus Number\n\n"
        "أرسل الرقم أو اكتب (تخطي/skip)\n"
        "Send the number or type (skip):"
    )
    return REG_FLIGHT


async def reg_flight(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text.lower() in ("skip", "تخطي"):
        context.user_data["reg"]["flight_number"] = ""
    else:
        context.user_data["reg"]["flight_number"] = text
    await update.message.reply_text(
        "📍 *6/8* مكان المغادرة (المطار/الميناء)\n"
        "Departure Location\n\n"
        "اختر من القائمة | Choose from the list:",
        reply_markup=LOCATION_KEYBOARD
    )
    return REG_LOCATION


async def reg_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data.startswith("loc:"):
        context.user_data["reg"]["departure_location"] = query.data[4:]
    await query.edit_message_text(
        f"📍 مكان المغادرة | Location: *{query.data[4:]}*\n─" * 15 + "\n\n"
        "🎫 *7/8* إرفاق صورة التذكرة (اختياري)\n"
        "Upload Flight Ticket (optional)\n\n"
        "أرسل صورة/PDF للتذكرة أو اكتب (تخطي/skip)\n"
        "Send ticket image/PDF or type (skip):"
    )
    return REG_FILE_TICKET


async def reg_file_ticket(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip().lower() if update.message.text else ""
    if text in ("skip", "تخطي"):
        context.user_data["reg"]["ticket_bytes"] = None
        context.user_data["reg"]["ticket_name"] = None
    elif update.message.document or update.message.photo:
        file = await (update.message.document if update.message.document else update.message.photo[-1]).get_file()
        context.user_data["reg"]["ticket_bytes"] = bytes(await file.download_as_bytearray())
        context.user_data["reg"]["ticket_name"] = (update.message.document.file_name if update.message.document else "ticket.jpg")
    else:
        context.user_data["reg"]["ticket_bytes"] = None
        context.user_data["reg"]["ticket_name"] = None
    await update.message.reply_text(
        "🛂 *8/8* إرفاق صورة الجواز أو بطاقة نسك (اختياري)\n"
        "Upload Passport Copy or Nusuk Card (optional)\n\n"
        "أرسل صورة/PDF أو اكتب (تخطي/skip)\n"
        "Send image/PDF or type (skip):"
    )
    return REG_FILE_PASSPORT


async def reg_file_passport(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip().lower() if update.message.text else ""
    if text in ("skip", "تخطي"):
        context.user_data["reg"]["passport_bytes"] = None
        context.user_data["reg"]["passport_name"] = None
    elif update.message.document or update.message.photo:
        file = await (update.message.document if update.message.document else update.message.photo[-1]).get_file()
        context.user_data["reg"]["passport_bytes"] = bytes(await file.download_as_bytearray())
        context.user_data["reg"]["passport_name"] = (update.message.document.file_name if update.message.document else "passport.jpg")
    else:
        context.user_data["reg"]["passport_bytes"] = None
        context.user_data["reg"]["passport_name"] = None

    reg = context.user_data["reg"]
    summary = (
        "📋 *ملخص الطلب | Request Summary*\n"
        f"─" * 20 + "\n"
        f"👤 الاسم | Name: *{reg.get('name', '—')}*\n"
        f"🛂 الجواز | Passport: *{reg.get('passport', '—')}*\n"
        f"📅 التاريخ | Date: *{reg.get('departure_date', '—') or '—'}*\n"
        f"⏰ الوقت | Time: *{reg.get('departure_time', '—') or '—'}*\n"
        f"✈️ الرحلة | Flight: *{reg.get('flight_number', '—') or '—'}*\n"
        f"📍 المكان | Location: *{reg.get('departure_location', '—') or '—'}*\n"
        f"🎫 التذكرة | Ticket: {'✅' if reg.get('ticket_bytes') else '—'}\n"
        f"🛂 الجواز | Passport: {'✅' if reg.get('passport_bytes') else '—'}\n"
        f"─" * 20 + "\n\n"
        "⚖️ *الإقرار | Declaration*\n\n"
        '"I hereby certify that all the information provided is accurate and matches my official travel documents, and I bear full responsibility for any incorrect data."\n\n'
        "اضغط للموافقة | Press to agree:",
    )
    await update.message.reply_text(summary, reply_markup=AGREE_KEYBOARD)
    return REG_DECLARATION


async def reg_declaration(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data != "agree:yes":
        await query.edit_message_text("❌ لم يتم الموافقة | Not agreed\nأرسل /register لإعادة المحاولة | Send /register to retry")
        return ConversationHandler.END

    reg = context.user_data["reg"]
    user_id = update.effective_user.id
    username = update.effective_user.username or ""

    tickets = {
        "passport": reg.get("passport", ""),
        "departure_date": reg.get("departure_date", ""),
        "departure_time": reg.get("departure_time", ""),
        "flight_number": reg.get("flight_number", ""),
        "departure_location": reg.get("departure_location", ""),
        "declaration": "agreed",
    }
    pilgrims = [{"name": reg.get("name", "")}]
    raw = ""

    try:
        eid = save_extraction(
            user_id=user_id,
            username=username,
            file_name=reg.get("ticket_name") or reg.get("passport_name") or "bot_manual",
            file_type="manual",
            pilgrims=pilgrims,
            tickets=tickets,
            raw_text=raw,
        )
    except Exception as e:
        await query.edit_message_text(f"❌ خطأ | Error: {e}")
        return ConversationHandler.END

    await query.edit_message_text(
        f"✅ *تم إرسال الطلب بنجاح!*\n"
        f"Request submitted successfully!\n\n"
        f"📋 *رقم الطلب | Request #: {eid}*\n"
        f"─" * 20 + "\n\n"
        "📌 سنتواصل معك قريباً\n"
        "We will contact you soon.\n\n"
        "/register لطلب جديد | New request\n"
        "/start للقائمة الرئيسية | Main menu"
    )
    return ConversationHandler.END


async def reg_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("❌ تم الإلغاء | Cancelled\n/register لبدء طلب جديد")
    return ConversationHandler.END


async def reg_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("❌ إدخال غير صحيح | Invalid input\nأرسل /register لبدء طلب جديد")
    return ConversationHandler.END


def main() -> None:
    if not TOKEN:
        logger.error("BOT_TOKEN not set")
        print("❌ BOT_TOKEN missing in .env")
        return

    app = Application.builder().token(TOKEN).build()

    reg_conv = ConversationHandler(
        entry_points=[CommandHandler("register", register_start)],
        states={
            REG_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)],
            REG_PASSPORT: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_passport)],
            REG_DEP_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_dep_date)],
            REG_DEP_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_dep_time)],
            REG_FLIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_flight)],
            REG_LOCATION: [CallbackQueryHandler(reg_location, pattern="^loc:")],
            REG_FILE_TICKET: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, reg_file_ticket),
                MessageHandler(filters.Document.ALL | filters.PHOTO, reg_file_ticket),
            ],
            REG_FILE_PASSPORT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, reg_file_passport),
                MessageHandler(filters.Document.ALL | filters.PHOTO, reg_file_passport),
            ],
            REG_DECLARATION: [CallbackQueryHandler(reg_declaration, pattern="^agree:")],
        },
        fallbacks=[CommandHandler("cancel", reg_cancel), MessageHandler(filters.COMMAND, reg_fallback)],
        per_message=False,
    )

    app.add_handler(reg_conv)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_handler))
    app.add_handler(CommandHandler("exportall", exportall_handler))
    app.add_handler(CommandHandler("history", history_handler))
    app.add_handler(CommandHandler("stats", stats_handler))
    app.add_handler(CommandHandler("sheets", sheets_handler))
    app.add_handler(MessageHandler(filters.Document.ALL & ~filters.COMMAND, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO & ~filters.COMMAND, handle_photo))

    print("✅ Bot running / البوت يعمل")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
