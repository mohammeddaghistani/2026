import sqlite3
import json
from datetime import datetime
from pathlib import Path

DB_PATH = Path("downloads") / "data.db"
EXCEL_PATH = Path("downloads") / "All_Pilgrims.xlsx"


def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS pilgrims_registry (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            passport TEXT UNIQUE NOT NULL,
            unified_permit TEXT,
            gender TEXT,
            group_visa TEXT,
            arrival_status TEXT,
            border_number TEXT,
            visa_number TEXT,
            hospitality_center TEXT,
            date_of_birth TEXT,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_registry_passport ON pilgrims_registry(passport);
        CREATE TABLE IF NOT EXISTS extractions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT,
            file_name TEXT,
            file_type TEXT,
            pilgrims TEXT,
            tickets TEXT,
            raw_text TEXT,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS pilgrims (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            extraction_id INTEGER,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            flight_number TEXT,
            ticket_number TEXT,
            seat TEXT,
            airline TEXT,
            passport TEXT,
            date TEXT,
            gate TEXT,
            booking TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (extraction_id) REFERENCES extractions(id)
        );
        CREATE INDEX IF NOT EXISTS idx_pilgrims_user ON pilgrims(user_id);
        CREATE INDEX IF NOT EXISTS idx_pilgrims_name ON pilgrims(name);
        CREATE INDEX IF NOT EXISTS idx_pilgrims_flight ON pilgrims(flight_number);
    """)
    for col in ["nationality", "departure_time", "departure_location", "declaration", "status", "status_updated_by", "status_updated_at"]:
        try:
            conn.execute(f"ALTER TABLE pilgrims ADD COLUMN {col} TEXT DEFAULT ''")
        except Exception:
            pass
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'organizer',
            created_at TEXT NOT NULL
        );
    """)
    conn.commit()
    conn.close()


def save_extraction(
    user_id: int, username: str, file_name: str, file_type: str,
    pilgrims: list, tickets: dict, raw_text: str
) -> int:
    conn = get_db()
    now = datetime.utcnow().isoformat()
    cur = conn.execute(
        "INSERT INTO extractions (user_id, username, file_name, file_type, pilgrims, tickets, raw_text, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, username, file_name, file_type,
         json.dumps(pilgrims, ensure_ascii=False),
         json.dumps(tickets, ensure_ascii=False),
         raw_text, now)
    )
    eid = cur.lastrowid

    flight = tickets.get("flight_number", "")
    tkt = tickets.get("ticket_number", "")
    for p in pilgrims:
        conn.execute(
            "INSERT INTO pilgrims (extraction_id, user_id, name, flight_number, ticket_number, seat, airline, passport, date, gate, booking, nationality, departure_time, departure_location, declaration, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (eid, user_id, p.get("name", ""),
             flight, tkt,
             tickets.get("seat", ""),
             tickets.get("airline", ""),
             tickets.get("passport", ""),
             tickets.get("date", ""),
             tickets.get("gate", ""),
             tickets.get("booking", ""),
             tickets.get("nationality", ""),
             tickets.get("departure_time", ""),
             tickets.get("departure_location", ""),
             tickets.get("declaration", ""),
             now)
        )
    conn.commit()
    conn.close()
    return eid


def get_history(user_id: int, limit: int = 0) -> list:
    conn = get_db()
    if limit:
        rows = conn.execute(
            "SELECT * FROM extractions WHERE user_id = ? ORDER BY created_at DESC LIMIT ?",
            (user_id, limit)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM extractions WHERE user_id = ? ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()
    conn.close()
    result = []
    for r in rows:
        result.append({
            "id": r["id"],
            "file_name": r["file_name"],
            "file_type": r["file_type"],
            "pilgrims": json.loads(r["pilgrims"]) if r["pilgrims"] else [],
            "tickets": json.loads(r["tickets"]) if r["tickets"] else {},
            "created_at": r["created_at"],
        })
    return result


def search_pilgrims(user_id: int, query: str = "") -> list:
    conn = get_db()
    if query:
        rows = conn.execute(
            "SELECT DISTINCT p.* FROM pilgrims p WHERE p.user_id = ? AND p.name LIKE ?",
            (user_id, f"%{query}%")
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM pilgrims WHERE user_id = ? ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_stats(user_id: int) -> dict:
    conn = get_db()
    total = conn.execute(
        "SELECT COUNT(*) as c FROM pilgrims WHERE user_id = ?", (user_id,)
    ).fetchone()["c"]
    flights = conn.execute(
        "SELECT COUNT(DISTINCT flight_number) as c FROM pilgrims WHERE user_id = ? AND flight_number != ''",
        (user_id,)
    ).fetchone()["c"]
    airlines = conn.execute(
        "SELECT COUNT(DISTINCT airline) as c FROM pilgrims WHERE user_id = ? AND airline != ''",
        (user_id,)
    ).fetchone()["c"]
    files = conn.execute(
        "SELECT COUNT(*) as c FROM extractions WHERE user_id = ?", (user_id,)
    ).fetchone()["c"]
    conn.close()
    return {
        "total_pilgrims": total,
        "total_flights": flights,
        "total_airlines": airlines,
        "total_files": files,
    }


def get_pilgrims_by_flight(user_id: int) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT flight_number, COUNT(*) as count, GROUP_CONCAT(name, ', ') as names "
        "FROM pilgrims WHERE user_id = ? AND flight_number != '' "
        "GROUP BY flight_number ORDER BY count DESC",
        (user_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


STATUS_OPTIONS = [
    ("processing", "جاري المعالجة | Processing"),
    ("under_action", "تحت الإجراء | Under Action"),
    ("completed", "تم إكمال اللازم | Completed"),
    ("departed", "الحاج مغادر | Departed"),
]

def update_status(pid: int, status: str, updated_by: str) -> bool:
    conn = get_db()
    now = datetime.utcnow().isoformat()
    cur = conn.execute(
        "UPDATE pilgrims SET status = ?, status_updated_by = ?, status_updated_at = ? WHERE id = ?",
        (status, updated_by, now, pid)
    )
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok

def get_pilgrim(pid: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM pilgrims WHERE id = ?", (pid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def passport_exists(passport: str) -> bool:
    conn = get_db()
    row = conn.execute(
        "SELECT 1 FROM pilgrims WHERE passport = ? LIMIT 1", (passport,)
    ).fetchone()
    conn.close()
    return row is not None


def get_all_pilgrims(search: str = "", limit: int = 500) -> list:
    conn = get_db()
    if search:
        rows = conn.execute(
            "SELECT * FROM pilgrims WHERE name LIKE ? OR passport LIKE ? OR flight_number LIKE ? ORDER BY created_at DESC LIMIT ?",
            (f"%{search}%", f"%{search}%", f"%{search}%", limit)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM pilgrims ORDER BY created_at DESC LIMIT ?", (limit,)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_users() -> list:
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, created_at FROM users ORDER BY created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def create_user(username: str, password: str, role: str = "organizer") -> int:
    conn = get_db()
    now = datetime.utcnow().isoformat()
    try:
        cur = conn.execute(
            "INSERT INTO users (username, password, role, created_at) VALUES (?, ?, ?, ?)",
            (username, password, role, now)
        )
        conn.commit()
        uid = cur.lastrowid
    except Exception as e:
        uid = None
    conn.close()
    return uid


def verify_user(username: str, password: str):
    conn = get_db()
    row = conn.execute(
        "SELECT id, username, role FROM users WHERE username = ? AND password = ?",
        (username, password)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_user(uid: int) -> bool:
    conn = get_db()
    cur = conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    conn.commit()
    deleted = cur.rowcount > 0
    conn.close()
    return deleted


def count_users() -> int:
    conn = get_db()
    row = conn.execute("SELECT COUNT(*) as c FROM users").fetchone()
    conn.close()
    return row["c"] if row else 0


def lookup_passport_registry(passport: str):
    conn = get_db()
    row = conn.execute("SELECT * FROM pilgrims_registry WHERE passport = ?", (passport.strip().upper(),)).fetchone()
    conn.close()
    return dict(row) if row else None


def import_excel_registry(excel_path: str = None) -> dict:
    """Import All_Pilgrims.xlsx into pilgrims_registry table. Returns stats."""
    import openpyxl
    path = Path(excel_path or EXCEL_PATH)
    if not path.exists():
        return {"error": f"File not found: {path}"}
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    col_count = len(ws[1])
    conn = get_db()
    added, skipped, errors = 0, 0, 0
    total_before = conn.execute("SELECT COUNT(*) FROM pilgrims_registry").fetchone()[0]
    for row in rows:
        if not row or not row[8] or not str(row[8]).strip():
            skipped += 1
            continue
        passport = str(row[8]).strip().upper()
        name = str(row[0] or "").strip() if col_count > 0 else ""
        unified_permit = str(row[1] or "").strip() if col_count > 1 else ""
        gender = str(row[2] or "").strip() if col_count > 2 else ""
        group_visa = str(row[3] or "").strip() if col_count > 3 else ""
        arrival_status = str(row[4] or "").strip() if col_count > 4 else ""
        border_number = str(row[5] or "").strip() if col_count > 5 else ""
        visa_number = str(row[6] or "").strip() if col_count > 6 else ""
        hospitality_center = str(row[7] or "").strip() if col_count > 7 else ""
        date_of_birth = str(row[9] or "").strip() if col_count > 9 else ""
        now = datetime.utcnow().isoformat()
        try:
            conn.execute(
                "INSERT OR IGNORE INTO pilgrims_registry (name, passport, unified_permit, gender, group_visa, arrival_status, border_number, visa_number, hospitality_center, date_of_birth, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (name, passport, unified_permit, gender, group_visa, arrival_status, border_number, visa_number, hospitality_center, date_of_birth, now)
            )
            added += 1
        except Exception:
            errors += 1
    conn.commit()
    total_after = conn.execute("SELECT COUNT(*) FROM pilgrims_registry").fetchone()[0]
    conn.close()
    return {"added": added, "skipped": skipped, "errors": errors, "total_in_excel": len(rows), "total_in_db": total_after, "new_records": total_after - total_before}


def count_registry() -> int:
    conn = get_db()
    row = conn.execute("SELECT COUNT(*) as c FROM pilgrims_registry").fetchone()
    conn.close()
    return row["c"] if row else 0


def get_pilgrims_by_airline(user_id: int) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT airline, COUNT(*) as count, GROUP_CONCAT(name, ', ') as names "
        "FROM pilgrims WHERE user_id = ? AND airline != '' "
        "GROUP BY airline ORDER BY count DESC",
        (user_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]
